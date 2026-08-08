"""
FileConnector — handles file uploads and normalises them into Document objects.

Supported file types:
- PDF (.pdf)         → extracted via pypdf
- Excel (.xlsx, .xls) → extracted via openpyxl (all sheets, row by row)
- CSV (.csv)         → extracted via stdlib csv module

All source-specific logic is contained in this class. The RAG pipeline
receives only ``Document`` objects and is completely unaware of file formats.

Configuration (via environment variables):
    MAX_FILE_SIZE_MB: Maximum allowed file size in megabytes (default: 50).
"""
import csv
import io
import logging
import os
from typing import Any, BinaryIO, List, Union

from connectors.base import BaseConnector
from connectors.document import Document
from connectors.exceptions import ConnectorError

logger = logging.getLogger(__name__)

# Attempt to import optional dependencies at module level for patchability in tests.
# If not installed, a ConnectorError is raised at call time (not at import time).
try:
    import pypdf  # type: ignore[import]
    _PYPDF_AVAILABLE = True
except ImportError:
    pypdf = None  # type: ignore[assignment]
    _PYPDF_AVAILABLE = False

try:
    import openpyxl  # type: ignore[import]
    _OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False

# Supported file extensions and their canonical source_type labels
_SUPPORTED_TYPES = {
    ".pdf": "pdf",
    ".xlsx": "excel",
    ".xls": "excel",
    ".csv": "csv",
}

_MAX_FILE_SIZE_BYTES: int = int(os.getenv("MAX_FILE_SIZE_MB", "50")) * 1024 * 1024
_MIN_TEXT_LENGTH: int = 10  # Minimum characters to be considered non-empty


class FileConnector(BaseConnector):
    """
    Connector for user-uploaded files (PDF, Excel, CSV).

    Parameters are passed via keyword arguments to :meth:`ingest`.

    All parsing errors, size violations, type mismatches, and empty-content
    situations raise :class:`~connectors.exceptions.ConnectorError` with a
    clear, user-readable message.
    """

    def ingest(
        self,
        file_bytes: bytes,
        filename: str,
        user_id: str,
        **kwargs: Any,
    ) -> List[Document]:
        """
        Parse an uploaded file and return normalised Document objects.

        Parameters:
            file_bytes: Raw bytes of the uploaded file.
            filename: Original filename, used as the source identifier and
                      to determine file type.
            user_id: The authenticated user who uploaded the file.
            **kwargs: Ignored; present for forward compatibility.

        Returns:
            A list of :class:`~connectors.document.Document` objects.
            For most files this is a single Document. Multi-sheet Excel
            files produce one Document per sheet that contains text.

        Raises:
            ValueError: If ``user_id`` is missing.
            ConnectorError: If the file is too large, an unsupported type,
                corrupted/unparseable, or contains no extractable text.
        """
        self.validate_user_id(user_id)

        if not filename or not filename.strip():
            raise ConnectorError("Filename must not be empty.", source="(unknown)")

        # ── Size validation ───────────────────────────────────────────────────
        file_size = len(file_bytes)
        if file_size == 0:
            raise ConnectorError(
                "The uploaded file is empty (0 bytes). Please upload a file with content.",
                source=filename,
            )
        if file_size > _MAX_FILE_SIZE_BYTES:
            limit_mb = _MAX_FILE_SIZE_BYTES // (1024 * 1024)
            actual_mb = file_size / (1024 * 1024)
            raise ConnectorError(
                f"File is too large ({actual_mb:.1f} MB). "
                f"Maximum allowed size is {limit_mb} MB.",
                source=filename,
            )

        # ── File type validation ──────────────────────────────────────────────
        lower_name = filename.lower()
        extension = ""
        for ext in _SUPPORTED_TYPES:
            if lower_name.endswith(ext):
                extension = ext
                break

        if not extension:
            supported = ", ".join(sorted(_SUPPORTED_TYPES.keys()))
            raise ConnectorError(
                f"Unsupported file type for '{filename}'. "
                f"Supported types: {supported}.",
                source=filename,
            )

        source_type = _SUPPORTED_TYPES[extension]
        logger.info(
            "FileConnector.ingest: filename=%r size=%d bytes type=%s user_id=%s",
            filename,
            file_size,
            source_type,
            user_id,
        )

        # ── Dispatch to type-specific parser ──────────────────────────────────
        if source_type == "pdf":
            documents = self._parse_pdf(file_bytes, filename, user_id)
        elif source_type == "excel":
            documents = self._parse_excel(file_bytes, filename, user_id)
        elif source_type == "csv":
            documents = self._parse_csv(file_bytes, filename, user_id)
        else:
            # Should be unreachable given the type check above, but be safe
            raise ConnectorError(
                f"Internal error: no parser for type '{source_type}'.",
                source=filename,
            )

        return documents

    # ── Private parsers ───────────────────────────────────────────────────────

    def _parse_pdf(self, file_bytes: bytes, filename: str, user_id: str) -> List[Document]:
        """
        Extract text from a PDF file using pypdf.

        Parameters:
            file_bytes: Raw PDF bytes.
            filename: Original filename (used as source identifier).
            user_id: The owning user.

        Returns:
            A single :class:`~connectors.document.Document` with the full
            extracted text from all pages.

        Raises:
            ConnectorError: If the file cannot be parsed as a PDF or
                contains no extractable text.
        """
        if not _PYPDF_AVAILABLE or pypdf is None:
            raise ConnectorError(
                "PDF parsing is unavailable: pypdf is not installed. "
                "Run: pip install pypdf",
                source=filename,
            )

        try:
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        except pypdf.errors.PdfReadError as exc:
            logger.warning(
                "FileConnector: corrupted PDF: filename=%r user_id=%s error=%s",
                filename,
                user_id,
                exc,
            )
            raise ConnectorError(
                f"The PDF file '{filename}' appears to be corrupted or is not a valid PDF. "
                "Please check the file and try again.",
                source=filename,
            ) from exc
        except Exception as exc:
            logger.error(
                "FileConnector: unexpected PDF read error: filename=%r user_id=%s error=%s",
                filename,
                user_id,
                exc,
                exc_info=True,
            )
            raise ConnectorError(
                f"Failed to open '{filename}' as a PDF: {exc}",
                source=filename,
            ) from exc

        page_count = len(reader.pages)
        texts: List[str] = []

        for page_num, page in enumerate(reader.pages, 1):
            try:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    texts.append(page_text)
            except Exception as exc:
                # Log but continue — a single bad page shouldn't fail the whole document
                logger.warning(
                    "FileConnector: failed to extract text from page %d of %r: %s",
                    page_num,
                    filename,
                    exc,
                )

        full_text = "\n\n".join(texts).strip()

        if not full_text or len(full_text) < _MIN_TEXT_LENGTH:
            raise ConnectorError(
                f"'{filename}' contains no extractable text. "
                "The PDF may be image-based (scanned) or protected. "
                "Please use a text-based PDF.",
                source=filename,
            )

        logger.info(
            "FileConnector: PDF extracted: filename=%r pages=%d chars=%d",
            filename,
            page_count,
            len(full_text),
        )

        return [
            Document(
                text=full_text,
                source_type="pdf",
                source_identifier=filename,
                user_id=user_id,
                metadata={
                    "page_count": page_count,
                    "filename": filename,
                },
            )
        ]

    def _parse_excel(
        self, file_bytes: bytes, filename: str, user_id: str
    ) -> List[Document]:
        """
        Extract text from an Excel file (.xlsx or .xls) using openpyxl.

        Each non-empty sheet produces a separate Document. This preserves
        sheet context so retrieval results can cite the specific sheet.

        Parameters:
            file_bytes: Raw Excel bytes.
            filename: Original filename (used as source identifier prefix).
            user_id: The owning user.

        Returns:
            A list of Documents — one per non-empty sheet.

        Raises:
            ConnectorError: If the file cannot be opened or all sheets are empty.
        """
        if not _OPENPYXL_AVAILABLE or openpyxl is None:
            raise ConnectorError(
                "Excel parsing is unavailable: openpyxl is not installed. "
                "Run: pip install openpyxl",
                source=filename,
            )

        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            logger.warning(
                "FileConnector: cannot open Excel file: filename=%r user_id=%s error=%s",
                filename,
                user_id,
                exc,
            )
            raise ConnectorError(
                f"'{filename}' could not be opened as an Excel file. "
                "The file may be corrupted or in an unsupported format.",
                source=filename,
            ) from exc

        documents: List[Document] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows: List[str] = []

            try:
                for row in sheet.iter_rows(values_only=True):
                    cell_values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                    if cell_values:
                        rows.append("\t".join(cell_values))
            except Exception as exc:
                logger.warning(
                    "FileConnector: error reading sheet %r in %r: %s",
                    sheet_name,
                    filename,
                    exc,
                )
                continue

            sheet_text = "\n".join(rows).strip()
            if not sheet_text or len(sheet_text) < _MIN_TEXT_LENGTH:
                logger.debug(
                    "FileConnector: skipping empty sheet %r in %r", sheet_name, filename
                )
                continue

            documents.append(
                Document(
                    text=sheet_text,
                    source_type="excel",
                    source_identifier=f"{filename}::{sheet_name}",
                    user_id=user_id,
                    metadata={
                        "filename": filename,
                        "sheet_name": sheet_name,
                        "row_count": len(rows),
                    },
                )
            )

        workbook.close()

        if not documents:
            raise ConnectorError(
                f"'{filename}' contains no extractable text across any sheet. "
                "Please ensure the file has data in at least one sheet.",
                source=filename,
            )

        logger.info(
            "FileConnector: Excel extracted: filename=%r sheets_with_data=%d",
            filename,
            len(documents),
        )
        return documents

    def _parse_csv(
        self, file_bytes: bytes, filename: str, user_id: str
    ) -> List[Document]:
        """
        Extract text from a CSV file using the stdlib ``csv`` module.

        Parameters:
            file_bytes: Raw CSV bytes.
            filename: Original filename.
            user_id: The owning user.

        Returns:
            A single :class:`~connectors.document.Document`.

        Raises:
            ConnectorError: If the CSV cannot be decoded or contains no data.
        """
        try:
            text_content = file_bytes.decode("utf-8", errors="replace")
        except Exception as exc:
            raise ConnectorError(
                f"'{filename}' could not be decoded as a CSV file.",
                source=filename,
            ) from exc

        try:
            reader = csv.reader(io.StringIO(text_content))
            rows: List[str] = []
            for row in reader:
                non_empty = [cell.strip() for cell in row if cell.strip()]
                if non_empty:
                    rows.append(", ".join(non_empty))
        except csv.Error as exc:
            logger.warning(
                "FileConnector: CSV parse error: filename=%r user_id=%s error=%s",
                filename,
                user_id,
                exc,
            )
            raise ConnectorError(
                f"'{filename}' could not be parsed as a CSV file: {exc}",
                source=filename,
            ) from exc

        full_text = "\n".join(rows).strip()
        if not full_text or len(full_text) < _MIN_TEXT_LENGTH:
            raise ConnectorError(
                f"'{filename}' contains no extractable data rows.",
                source=filename,
            )

        logger.info(
            "FileConnector: CSV extracted: filename=%r rows=%d chars=%d",
            filename,
            len(rows),
            len(full_text),
        )

        return [
            Document(
                text=full_text,
                source_type="csv",
                source_identifier=filename,
                user_id=user_id,
                metadata={
                    "filename": filename,
                    "row_count": len(rows),
                },
            )
        ]
